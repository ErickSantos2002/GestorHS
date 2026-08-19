"""Motor de planilhas do sistema.

Puro: nao conhece SQLAlchemy, FastAPI nem o dominio. Recebe colunas e dicionarios,
devolve os bytes de um .xlsx. Existe para que TODA exportacao do GestorHS saia com o
mesmo acabamento — se o cabecalho muda aqui, muda em todas de uma vez.
"""
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Literal, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

Formato = Literal["texto", "data", "datahora", "numero", "inteiro", "sim_nao"]

# Teto de seguranca. Acima disso a geracao sincrona comeca a segurar o worker, e a
# planilha ja passou do tamanho em que alguem consegue trabalhar nela.
LIMITE_LINHAS = 50_000

_FORMATO_NUMERICO = {
    "data": "DD/MM/YYYY",
    "datahora": "DD/MM/YYYY HH:MM",
    "numero": "#,##0.00",
    "inteiro": "0",
}

_FUNDO_CABECALHO = PatternFill("solid", start_color="FF1F3B57", end_color="FF1F3B57")
_FONTE_CABECALHO = Font(bold=True, color="FFFFFFFF")


class PlanilhaGrandeDemais(Exception):
    """Mais linhas do que LIMITE_LINHAS. O chamador transforma em 400."""


@dataclass(frozen=True)
class Coluna:
    titulo: str
    campo: str
    largura: int
    formato: Formato = "texto"


def _valor_da_celula(bruto: Any, formato: Formato) -> Any:
    if bruto is None:
        # Celula VAZIA, nunca o "—" que a tela usa: quem recebe a planilha filtra e
        # soma em cima dela, e um travessao transforma a coluna inteira em texto.
        return None
    if formato == "sim_nao":
        return "Sim" if bruto else "Nao"
    if formato == "datahora" and isinstance(bruto, datetime):
        # O Excel nao tem conceito de fuso e o openpyxl recusa datetime aware.
        # As datas do sistema sao UTC; guardamos o instante sem o rotulo.
        return bruto.replace(tzinfo=None)
    if formato == "data" and isinstance(bruto, datetime):
        return bruto.date()
    if formato in ("numero", "inteiro") and isinstance(bruto, Decimal):
        return float(bruto)
    return bruto


def gerar_xlsx(
    titulo_aba: str,
    colunas: Sequence[Coluna],
    linhas: Sequence[dict],
    rodape: str,
) -> bytes:
    """Monta a planilha e devolve os bytes. `linhas` sao dicionarios; a chave usada
    de cada uma e' o `campo` da Coluna, e chave ausente vale o mesmo que None."""
    if len(linhas) > LIMITE_LINHAS:
        raise PlanilhaGrandeDemais(len(linhas))

    wb = Workbook()
    aba = wb.active
    aba.title = titulo_aba[:31]  # limite do proprio formato xlsx

    aba.append([c.titulo for c in colunas])
    for i, coluna in enumerate(colunas, start=1):
        celula = aba.cell(row=1, column=i)
        celula.font = _FONTE_CABECALHO
        celula.fill = _FUNDO_CABECALHO
        celula.alignment = Alignment(vertical="center")
        aba.column_dimensions[get_column_letter(i)].width = coluna.largura

    for linha in linhas:
        aba.append([_valor_da_celula(linha.get(c.campo), c.formato) for c in colunas])

    for i, coluna in enumerate(colunas, start=1):
        formato_numerico = _FORMATO_NUMERICO.get(coluna.formato)
        if not formato_numerico:
            continue
        for celula in aba.iter_rows(min_row=2, max_row=1 + len(linhas),
                                    min_col=i, max_col=i):
            celula[0].number_format = formato_numerico

    # Congela o cabecalho e liga o autofiltro em toda a faixa de dados. Com o
    # cabecalho fixo da' para rolar mil linhas sem perder de vista o que e' cada coluna.
    aba.freeze_panes = "A2"
    aba.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{1 + len(linhas)}"

    # Uma linha em branco separa os dados do rodape — sem ela o autofiltro do Excel
    # trataria o rodape como se fosse mais um registro.
    aba.append([])
    aba.append([rodape])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
