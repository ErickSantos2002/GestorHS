from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _headers(client, email="admin@hs.com", senha="senha123"):
    tok = client.post("/auth/login", json={"email": email, "senha": senha}).json()
    return {"Authorization": f"Bearer {tok['access_token']}"}


def _series(resposta):
    """Todas as celulas da planilha como texto, para procurar valores dentro dela."""
    aba = load_workbook(BytesIO(resposta.content)).active
    return {str(c.value) for linha in aba.iter_rows() for c in linha if c.value is not None}


def _base(db_session):
    from app.models import Cliente, Equipamento, Marca
    marca = Marca(descricao="Drager")
    db_session.add(marca)
    db_session.flush()
    c = Cliente(nome="Cliente Export", cgc="11222333000144")
    e = Equipamento(descricao="Alcotest 6820", marca=marca.id)
    db_session.add_all([c, e])
    db_session.commit()
    return c.id, e.id


def test_exportar_frota_exige_token(client):
    assert client.get("/equipamentos-cliente/exportar").status_code == 401


def test_exportar_frota_devolve_xlsx_com_nome_de_arquivo(client, usuario_admin, db_session):
    _base(db_session)
    r = client.get("/equipamentos-cliente/exportar", headers=_headers(client))
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX)
    assert "attachment" in r.headers["content-disposition"]
    assert f"equipamentos-{date.today().isoformat()}.xlsx" in r.headers["content-disposition"]


def test_exportar_frota_respeita_o_filtro_de_status(client, usuario_admin, db_session):
    """O teste que importa: a planilha nao pode trazer linha que o filtro exclui."""
    from app.models import EquipamentoCliente
    cid, eid = _base(db_session)
    hoje = date.today()
    db_session.add_all([
        EquipamentoCliente(cliente=cid, equipamento=eid, serie="SERIEVENCIDA",
                           prox_calibragem=hoje - timedelta(days=1)),
        EquipamentoCliente(cliente=cid, equipamento=eid, serie="SERIEEMDIA",
                           prox_calibragem=hoje + timedelta(days=200)),
    ])
    db_session.commit()
    r = client.get("/equipamentos-cliente/exportar?status=vencido", headers=_headers(client))
    valores = _series(r)
    assert "SERIEVENCIDA" in valores
    assert "SERIEEMDIA" not in valores


def test_exportar_frota_ignora_a_paginacao_da_tela(client, usuario_admin, db_session):
    """A tela mostra 25 por vez; a planilha tem que trazer tudo."""
    from app.models import EquipamentoCliente
    cid, eid = _base(db_session)
    db_session.add_all([
        EquipamentoCliente(cliente=cid, equipamento=eid, serie=f"S{i:03d}")
        for i in range(40)
    ])
    db_session.commit()
    valores = _series(client.get("/equipamentos-cliente/exportar", headers=_headers(client)))
    assert "S000" in valores and "S039" in valores


def test_exportar_frota_traz_a_marca_e_o_cnpj_do_cliente(client, usuario_admin, db_session):
    from app.models import EquipamentoCliente
    cid, eid = _base(db_session)
    db_session.add(EquipamentoCliente(cliente=cid, equipamento=eid, serie="COMMARCA"))
    db_session.commit()
    valores = _series(client.get("/equipamentos-cliente/exportar", headers=_headers(client)))
    assert "Drager" in valores
    assert "11222333000144" in valores


def test_exportar_clientes_exige_token(client):
    assert client.get("/clientes/exportar").status_code == 401


def test_exportar_clientes_respeita_a_busca(client, usuario_admin, db_session):
    from app.models import Cliente
    db_session.add_all([Cliente(nome="Alfa Industria"), Cliente(nome="Beta Comercio")])
    db_session.commit()
    valores = _series(client.get("/clientes/exportar?q=Alfa", headers=_headers(client)))
    assert "Alfa Industria" in valores
    assert "Beta Comercio" not in valores


def test_exportar_frota_filtrando_por_cliente_mostra_o_nome_no_rodape(client, usuario_admin, db_session):
    """O rodape existe pra dizer com que filtro a planilha saiu — mostrar so' o id
    do cliente nao ajuda quem recebe o arquivo por e-mail."""
    from app.models import EquipamentoCliente
    cid, eid = _base(db_session)
    db_session.add(EquipamentoCliente(cliente=cid, equipamento=eid, serie="COMFILTRO"))
    db_session.commit()
    valores = _series(client.get(f"/equipamentos-cliente/exportar?cliente={cid}", headers=_headers(client)))
    assert any("Cliente Export" in v for v in valores)
    assert not any(v == f"Cliente: {cid}" for v in valores)


def test_exportar_clientes_nao_colide_com_a_rota_de_id(client, usuario_admin, db_session):
    """Se /exportar for declarado depois de /{cliente_id}, o FastAPI tenta converter
    "exportar" para int e devolve 422 em vez do arquivo."""
    r = client.get("/clientes/exportar", headers=_headers(client))
    assert r.status_code == 200, r.text


def test_exportar_ordens_exige_token(client):
    assert client.get("/ordens/exportar").status_code == 401


def test_exportar_ordens_respeita_o_filtro_de_fase(client, usuario_admin, db_session, fases_seed):
    from app.models import Ordem
    cid, eid = _base(db_session)
    db_session.add_all([
        Ordem(cliente=cid, fase=4, etiqueta="ETIQ-RECEBIDO"),
        Ordem(cliente=cid, fase=5, etiqueta="ETIQ-LAB"),
    ])
    db_session.commit()
    valores = _series(client.get("/ordens/exportar?fase=4", headers=_headers(client)))
    assert "ETIQ-RECEBIDO" in valores
    assert "ETIQ-LAB" not in valores


def test_exportar_ordens_nao_colide_com_a_rota_de_id(client, usuario_admin):
    assert client.get("/ordens/exportar", headers=_headers(client)).status_code == 200


def test_acima_do_teto_devolve_400(client, usuario_admin, db_session, monkeypatch):
    from app.models import EquipamentoCliente
    import app.core.planilha as planilha
    monkeypatch.setattr(planilha, "LIMITE_LINHAS", 1)
    cid, eid = _base(db_session)
    db_session.add_all([
        EquipamentoCliente(cliente=cid, equipamento=eid, serie="A"),
        EquipamentoCliente(cliente=cid, equipamento=eid, serie="B"),
    ])
    db_session.commit()
    r = client.get("/equipamentos-cliente/exportar", headers=_headers(client))
    assert r.status_code == 400
    assert "filtro" in r.json()["detail"].lower()
