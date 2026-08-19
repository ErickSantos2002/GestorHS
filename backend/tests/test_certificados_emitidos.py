from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import load_workbook


def _headers(client, email="admin@hs.com", senha="senha123"):
    tok = client.post("/auth/login", json={"email": email, "senha": senha}).json()
    return {"Authorization": f"Bearer {tok['access_token']}"}


def _valores(resposta):
    aba = load_workbook(BytesIO(resposta.content)).active
    return {str(c.value) for linha in aba.iter_rows() for c in linha if c.value is not None}


def _cenario(db_session):
    """Um certificado vindo de OS e um de venda, para o mesmo aparelho."""
    from app.models import (Cliente, Equipamento, EquipamentoCliente, Ordem,
                            OSCertificado, CertificadoVenda)
    c = Cliente(nome="Cliente Cert", cgc="99888777000166")
    e = Equipamento(descricao="Alcotest 6820")
    db_session.add_all([c, e])
    db_session.flush()
    ec = EquipamentoCliente(cliente=c.id, equipamento=e.id, serie="SERIE-CERT")
    db_session.add(ec)
    db_session.flush()
    o = Ordem(cliente=c.id, equipamento_cliente=ec.id, calib_cert="CERT-OS-1",
              data_calibracao=datetime(2026, 3, 10, tzinfo=timezone.utc))
    db_session.add(o)
    db_session.flush()
    db_session.add_all([
        OSCertificado(os=o.id, tipo="C",
                      data_geracao=datetime(2026, 3, 11, tzinfo=timezone.utc)),
        CertificadoVenda(equipamento_cliente=ec.id, html="<p>x</p>",
                         calib_cert="CERT-VENDA-1", data_calibracao=date(2025, 1, 5),
                         data_geracao=datetime(2025, 1, 6, tzinfo=timezone.utc)),
    ])
    db_session.commit()
    return c.id


def test_exige_token(client):
    assert client.get("/certificados-emitidos/exportar").status_code == 401


def test_une_certificados_de_os_e_de_venda(client, usuario_admin, db_session):
    _cenario(db_session)
    valores = _valores(client.get("/certificados-emitidos/exportar", headers=_headers(client)))
    assert "CERT-OS-1" in valores
    assert "CERT-VENDA-1" in valores
    assert "OS" in valores and "Venda" in valores


def test_filtro_de_periodo_corta_pela_data_de_geracao(client, usuario_admin, db_session):
    _cenario(db_session)
    r = client.get("/certificados-emitidos/exportar?de=2026-01-01", headers=_headers(client))
    valores = _valores(r)
    assert "CERT-OS-1" in valores
    assert "CERT-VENDA-1" not in valores


def test_filtro_de_cliente(client, usuario_admin, db_session):
    from app.models import Cliente
    cid = _cenario(db_session)
    outro = Cliente(nome="Outro Cliente")
    db_session.add(outro)
    db_session.commit()
    valores_outro = _valores(client.get(f"/certificados-emitidos/exportar?cliente={outro.id}",
                                        headers=_headers(client)))
    assert "CERT-OS-1" not in valores_outro
    assert "CERT-VENDA-1" not in valores_outro

    valores_certo = _valores(client.get(f"/certificados-emitidos/exportar?cliente={cid}",
                                        headers=_headers(client)))
    assert "CERT-OS-1" in valores_certo
    assert "CERT-VENDA-1" in valores_certo


def test_tipo_sai_por_extenso(client, usuario_admin, db_session):
    _cenario(db_session)
    valores = _valores(client.get("/certificados-emitidos/exportar", headers=_headers(client)))
    assert "Calibracao" in valores


def _cenario_periodo(db_session):
    """Dois certificados de OS para provar o fim de faixa inclusivo de `ate`.

    Um gerado as 14h do dia do filtro (tem que aparecer — e' o caso que o comentario
    do codigo descreve); outro gerado exatamente a meia-noite do dia seguinte (tem
    que ficar de fora).
    """
    from app.models import Cliente, Equipamento, EquipamentoCliente, Ordem, OSCertificado
    c = Cliente(nome="Cliente Periodo")
    e = Equipamento(descricao="Alcotest 7110")
    db_session.add_all([c, e])
    db_session.flush()
    ec = EquipamentoCliente(cliente=c.id, equipamento=e.id, serie="SERIE-PERIODO")
    db_session.add(ec)
    db_session.flush()
    dentro = Ordem(cliente=c.id, equipamento_cliente=ec.id, calib_cert="CERT-DENTRO-14H",
                    data_calibracao=datetime(2026, 3, 15, tzinfo=timezone.utc))
    fora = Ordem(cliente=c.id, equipamento_cliente=ec.id, calib_cert="CERT-FORA-DIA-SEGUINTE",
                 data_calibracao=datetime(2026, 3, 16, tzinfo=timezone.utc))
    db_session.add_all([dentro, fora])
    db_session.flush()
    db_session.add_all([
        OSCertificado(os=dentro.id, tipo="C",
                      data_geracao=datetime(2026, 3, 15, 14, 0, tzinfo=timezone.utc)),
        OSCertificado(os=fora.id, tipo="C",
                      data_geracao=datetime(2026, 3, 16, 0, 0, tzinfo=timezone.utc)),
    ])
    db_session.commit()


def test_filtro_ate_inclui_o_dia_inteiro_e_corta_no_seguinte(client, usuario_admin, db_session):
    _cenario_periodo(db_session)
    valores = _valores(client.get("/certificados-emitidos/exportar?ate=2026-03-15",
                                  headers=_headers(client)))
    assert "CERT-DENTRO-14H" in valores
    assert "CERT-FORA-DIA-SEGUINTE" not in valores
